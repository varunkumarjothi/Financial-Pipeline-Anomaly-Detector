"""
Financial Report Generator
===========================
Generates charts and an Excel report from the processed financial data.
Author: Varun Kumar Jothi
"""

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import matplotlib.patches as mpatches
import sqlite3
import os
import sys
import logging

log = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

BASE_DIR    = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB_PATH     = os.path.join(BASE_DIR, 'data', 'financial.db')
REPORTS_DIR = os.path.join(BASE_DIR, 'reports')
os.makedirs(REPORTS_DIR, exist_ok=True)

COLORS  = ['#1A5276','#E74C3C','#27AE60','#F39C12','#8E44AD']
SYMBOLS = ["RELIANCE.NS","TCS.NS","INFY.NS","HDFCBANK.NS","WIPRO.NS"]


def save_chart(fig, name):
    path = os.path.join(REPORTS_DIR, name)
    fig.savefig(path, dpi=150, bbox_inches='tight')
    plt.close(fig)
    log.info(f"  Chart: {path}")


def chart_price_history(df):
    fig, axes = plt.subplots(3, 2, figsize=(16, 14))
    axes = axes.flatten()
    for i, (symbol, grp) in enumerate(df.groupby('Symbol')):
        if i >= 5: break
        ax  = axes[i]
        grp = grp.sort_values('Date')
        ax.plot(grp['Date'], grp['Close'], color=COLORS[i], linewidth=1.5, label='Close')
        if 'ma_20' in grp.columns:
            ax.plot(grp['Date'], grp['ma_20'], color='orange', linewidth=1,
                    linestyle='--', alpha=0.8, label='MA20')
        if 'is_anomaly' in grp.columns:
            anoms = grp[grp['is_anomaly'] == True]
            if len(anoms):
                ax.scatter(anoms['Date'], anoms['Close'], color='red',
                           s=30, zorder=5, label='Anomaly')
        ax.set_title(f"{symbol}", fontsize=11, fontweight='bold')
        ax.set_ylabel('Price (Rs.)')
        ax.legend(fontsize=8)
        ax.grid(linestyle='--', alpha=0.4)
        ax.xaxis.set_major_locator(plt.MaxNLocator(6))
        plt.setp(ax.xaxis.get_majorticklabels(), rotation=30, ha='right', fontsize=8)
    axes[-1].axis('off')
    fig.suptitle('Stock Price History with Anomalies & MA20',
                 fontsize=15, fontweight='bold', y=1.01)
    plt.tight_layout()
    save_chart(fig, 'chart_01_price_history.png')


def chart_returns_distribution(df):
    fig, axes = plt.subplots(1, 5, figsize=(18, 5), sharey=False)
    for i, (symbol, grp) in enumerate(df.groupby('Symbol')):
        if i >= 5: break
        ret = grp['daily_return_%'].dropna()
        axes[i].hist(ret, bins=40, color=COLORS[i], alpha=0.75, edgecolor='white')
        axes[i].axvline(ret.mean(), color='black', linestyle='--',
                        linewidth=1.5, label=f'Mean: {ret.mean():.2f}%')
        axes[i].axvline(ret.mean()+2*ret.std(), color='red', linestyle=':', linewidth=1.2)
        axes[i].axvline(ret.mean()-2*ret.std(), color='red', linestyle=':',
                        linewidth=1.2, label='±2σ')
        axes[i].set_title(symbol.replace('.NS',''), fontsize=10, fontweight='bold')
        axes[i].set_xlabel('Daily Return %')
        axes[i].legend(fontsize=8)
        axes[i].grid(axis='y', linestyle='--', alpha=0.4)
    fig.suptitle('Daily Return Distribution by Stock', fontsize=14, fontweight='bold')
    plt.tight_layout()
    save_chart(fig, 'chart_02_returns_distribution.png')


def chart_anomaly_timeline(anomalies):
    if anomalies.empty:
        return
    anomalies = anomalies.copy()
    anomalies['Date'] = pd.to_datetime(anomalies['Date'])
    severity_colors = {
        'Critical': '#E74C3C',
        'High':     '#F39C12',
        'Moderate': '#F7DC6F',
        'Watch':    '#AED6F1'
    }

    fig, ax = plt.subplots(figsize=(16, 6))
    for sym_i, (symbol, grp) in enumerate(anomalies.groupby('Symbol')):
        for _, row in grp.iterrows():
            color = severity_colors.get(row['severity'], 'grey')
            ax.scatter(row['Date'], sym_i, color=color, s=60, zorder=5)
            if row['severity'] in ['Critical', 'High']:
                ax.annotate(f"{row['daily_return_%']:+.1f}%",
                            (row['Date'], sym_i),
                            textcoords="offset points",
                            xytext=(0, 10), fontsize=7,
                            ha='center', color=color)
    ax.set_yticks(range(anomalies['Symbol'].nunique()))
    ax.set_yticklabels(anomalies['Symbol'].unique(), fontsize=10)
    ax.set_title('Anomaly Timeline by Stock', fontsize=14, fontweight='bold', pad=12)
    ax.set_xlabel('Date')
    patches = [mpatches.Patch(color=v, label=k) for k, v in severity_colors.items()]
    ax.legend(handles=patches, loc='upper right')
    ax.grid(axis='x', linestyle='--', alpha=0.4)
    plt.tight_layout()
    save_chart(fig, 'chart_03_anomaly_timeline.png')


def chart_volatility_comparison(df):
    stats = df.groupby('Symbol')['volatility_20d'].mean().sort_values(ascending=False)
    fig, ax = plt.subplots(figsize=(10, 5))
    bars = ax.bar(stats.index, stats.values, color=COLORS[:len(stats)])
    for bar in bars:
        ax.text(bar.get_x()+bar.get_width()/2, bar.get_height()+0.02,
                f'{bar.get_height():.2f}%', ha='center', fontweight='bold')
    ax.set_title('Average 20-Day Volatility by Stock',
                 fontsize=14, fontweight='bold', pad=12)
    ax.set_ylabel('Volatility (%)')
    ax.grid(axis='y', linestyle='--', alpha=0.4)
    plt.tight_layout()
    save_chart(fig, 'chart_04_volatility.png')


def generate_excel(df, anomalies):
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment

    path  = os.path.join(REPORTS_DIR, 'financial_analysis_report.xlsx')
    stats = df.groupby('Symbol').agg(
        trading_days=('Close', 'count'),
        avg_close=('Close', 'mean'),
        min_close=('Close', 'min'),
        max_close=('Close', 'max'),
        avg_daily_return=('daily_return_%', 'mean'),
        avg_volatility=('volatility_20d', 'mean'),
        avg_volume_ratio=('volume_ratio', 'mean'),
    ).round(2).reset_index()

    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        stats.to_excel(writer, sheet_name='Symbol Statistics', index=False)
        if not anomalies.empty:
            disp_cols = ['Symbol','Company','Date','Close','daily_return_%',
                         'anomaly_score','severity','anomaly_direction']
            anomalies[[c for c in disp_cols if c in anomalies.columns]]\
                .to_excel(writer, sheet_name='Anomalies', index=False)
            anomalies.groupby(['Symbol','severity']).size()\
                .reset_index(name='count')\
                .to_excel(writer, sheet_name='Anomaly Summary', index=False)

        # Auto-fit columns and style headers
        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            for cell in ws[1]:
                cell.font      = Font(bold=True, color='FFFFFF')
                cell.fill      = PatternFill(start_color='1A5276',
                                             end_color='1A5276',
                                             fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            for col in ws.columns:
                max_len    = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max_len + 4

    log.info(f"  Excel report: {path}")


def run_report():
    log.info("="*55)
    log.info("  FINANCIAL REPORT GENERATOR")
    log.info("="*55)

    conn = sqlite3.connect(DB_PATH)
    df   = pd.read_sql("SELECT * FROM market_data", conn)
    try:
        anomalies = pd.read_sql("SELECT * FROM anomalies", conn)
    except Exception:
        anomalies = pd.DataFrame()
    conn.close()

    df['Date'] = pd.to_datetime(df['Date'])
    if not anomalies.empty:
        anomalies['Date'] = pd.to_datetime(anomalies['Date'])
        df = df.merge(
            anomalies[['Symbol','Date','anomaly_score','severity']],
            on=['Symbol','Date'], how='left'
        )
        df['is_anomaly'] = df['anomaly_score'].fillna(0) >= 2

    log.info("\nGenerating charts...")
    chart_price_history(df)
    chart_returns_distribution(df)
    chart_anomaly_timeline(anomalies)
    chart_volatility_comparison(df)

    log.info("\nGenerating Excel report...")
    generate_excel(df, anomalies)

    log.info("\n  Report generation complete!")


if __name__ == "__main__":
    run_report()